#!/usr/bin/env python3
from pyVim.connect import SmartConnect, Disconnect
from pyVmomi import vim, vmodl
from getpass import getpass
import ssl
import atexit
import argparse
import openpyxl
import re
__version__ = "0.0.1"

class login:
    """
    ログイン処理を抽象化したクラス
    """
    def __init__(self):
        self.username = ""
        self.password = ""
        self.host = ""

    def get_service_instance(self):
        """
        ServiceInstanceを取得するメソッド

        :rtype: class
        :return: pyVmomi.VmomiSupport.vim.ServiceInstance
        """
        # SSL証明書対策
        context = None
        if hasattr(ssl, '_create_unverified_context'):
            context = ssl._create_unverified_context()

        # ServiceInstanceを取得
        si = SmartConnect(host = self.host,
                          user = self.username,
                          pwd = self.password,
                          sslContext = context)

        # 処理完了時にvCenterから切断
        atexit.register(Disconnect, si)

        return si

def options():
    """
    コマンドラインオプション設定

    :rtype: class
    :return: argparse.Namespace
    """
    parser = argparse.ArgumentParser(prog='',
                                     add_help=True,
                                     description='')
    parser.add_argument('--version', '-v',
                        action='version',
                        version=__version__)
    parser.add_argument('--host', '-vc',
                        type=str, required=True,
                        help='vCenterのIP又はホスト名')
    parser.add_argument('--username', '-u',
                        type=str, default='administrator@vsphere.local',
                        help='vCenterのログインユーザー名(default:administrator@vsphere.local)')
    parser.add_argument('--password', '-p',
                        type=str,
                        help='vCenterのログインユーザーパスワード')
    parser.add_argument('--output',
                        type=str, default='output.xlsx',
                        help='出力するExcelファイル名(defualt:output.xlsx)')
    args = parser.parse_args()

    if(not(args.password)):
        args.password = getpass()

    return args

def get_vmware_mob(si, mob):
    """
    Managed Objectの配列を取得して返す

    :type si: class
    :param si: ServiceInstanceオブジェクト

    :type mob: class
    :param mob: Managed Object class
    :return: オブジェクトのリスト
    """
    # HostSystemのオブジェクトを取得
    mob_list = si.content.viewManager.CreateContainerView(si.content.rootFolder,
                                                          [mob],
                                                          True)

    return mob_list

def create_esxi_inventorys(mob_list, ws):
    """
    ESXiのインベントリファイルを作成する関数

    :type mob_list: list
    :param mob_list: ホストオブジェクトのリスト

    :type ws: class
    :param ws: ワークシートのオブジェクト
    """
    # Inventoryのカラム名
    column_name = [
        "HostName",
        "Vendor",
        "Model",
        "uuid",
        "BIOS Version",
        "CPU Model",
        "CPU Hz",
        "CPU Socket",
        "CPU Core",
        "CPU Thread",
        "CPU HyperThread",
        "Memory Size",
        "ESXi Version",
        "ESXi Build Version",
        "ESXi API Version",
        "ESXi Cluster",
        "ESXi Datastore",
        "ESXi VM",
        "ESXi ManageIP",
        "ESXi SubnetMask",
        "ESXi MacAddress",
        "ESXi ManagePG",
        "ESXi DefaultGW",
        "ESXi DNS"
    ]

    all_inventorys = []
    for host in mob_list.view:
        inventorys = []
        # 基本情報
        inventorys.append(host.name)
        inventorys.append(host.hardware.systemInfo.vendor)      # サーバーベンダー
        inventorys.append(host.hardware.systemInfo.model)       # サーバーモデル
        inventorys.append(host.hardware.systemInfo.uuid)                   # サーバuuid
        inventorys.append(host.hardware.biosInfo.biosVersion)   # サーバBIOSバージョン

        # CPU情報
        inventorys.append(host.hardware.cpuPkg[0].description)  # CPUモデル
        inventorys.append(host.hardware.cpuPkg[0].hz)           # CPU周波数
        inventorys.append(host.hardware.cpuInfo.numCpuPackages) # CPUソケット
        inventorys.append(host.hardware.cpuInfo.numCpuCores)    # CPUコア数
        inventorys.append(host.hardware.cpuInfo.numCpuThreads)  # CPUスレッド数
        inventorys.append(host.config.hyperThread.available)    # CPUのハイパースレッド

        # メモリー情報
        inventorys.append(host.hardware.memorySize)             # メモリーサイズ

        # ESXi情報
        inventorys.append(host.summary.config.product.version)    # ESXiバージョン
        inventorys.append(host.summary.config.product.build)      # ESXiビルドバージョン
        inventorys.append(host.summary.config.product.apiVersion) # ESXi APIバージョン
        inventorys.append((lambda  x: x.name if(isinstance(x, vim.ClusterComputeResource)) else "")(host.parent)) # 所属クラスタ
        inventorys.append(",".join(sorted(list(map(lambda x: x.name, host.datastore))))) # ESXiがマウントしてるデータストア
        inventorys.append(len(host.vm))                         # ESXi上のVM数(テンプレート含む)

        # ESXi管理IP関連情報
        mngIp = (list(filter(lambda x: x.device == "vmk0", host.config.network.vnic))[0])
        inventorys.append(mngIp.spec.ip.ipAddress)              # ESXiの管理IP
        inventorys.append(mngIp.spec.ip.subnetMask)             # ESXiの管理IPサブネットマスク
        inventorys.append(mngIp.spec.mac)                       # ESXiの管理IPデバイスのMacアドレス
        inventorys.append(mngIp.spec.portgroup)                 # ESXiの管理IPのポートグループ
        inventorys.append(mngIp.spec.ipRouteSpec.ipRouteConfig.defaultGateway) # ESXiの管理IPデフォルトGW
        inventorys.append(",".join(sorted(host.config.network.dnsConfig.address))) # ESXiに設定されているDNS

        # Inventoryを追加
        (lambda x: all_inventorys.append(x))(inventorys)


    # WSへInventory情報追加
    ws.append(column_name)
    for esxi_inventory in sorted(all_inventorys):
        ws.append(esxi_inventory)

    # Inventoryカラム名セルの塗りつぶし
    fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='629BF7')
    for row in ws:
        # ソート設定
        s = row[0]
        e = row[len(row) - 1]
        ws.auto_filter.ref = "%s%s:%s%s" % (s.column, s.col_idx, e.column, e.col_idx)

        # セルの塗りつぶし
        for cell in row:
            cell.fill = fill
        break

    # WSタブの色設定
    ws.sheet_properties.tabColor = "1072BA"

def create_vm_inventorys(mob_list, dp_dict, ws):
    """
    VMのインベントリファイルを作成する関数

    :type mob_list: list
    :param mob_list: 仮想マシンオブジェクトのリスト

    :type dp_dict: dict
    :param dp_dict: 分散ポートグループのキーからポートグループ名を取得できる辞書

    :type ws: class
    :param ws: ワークシートのオブジェクト
    """
    # Inventoryカラム名
    column_name = [
        'HostName',
        'InstanceUuid',
        "ResourcePool",
        "VMwareTools Status",
        "VMwareTools Version",
        'OS',
        "CPU Socket",
        "CPU Core",
        "CPU Reservation",
        "CPU Limit",
        "Memory Size",
        "Memory Reservation",
        "Memory Limit",
        "CDROM",
        "Floppy",
        "USB",
        "Disk Total Size",
        "IPAddress",
        "Network Adapter 1 PG",
        "Network Adapter 2 PG",
        "Network Adapter 3 PG",
        "Network Adapter 4 PG",
        "Network Adapter 5 PG",
        "Network Adapter 6 PG",
        "Network Adapter 7 PG",
        "Network Adapter 8 PG",
        "Network Adapter 9 PG",
        "Network Adapter 10 PG",
    ]

    all_inventorys = []
    for vm in mob_list.view:
        inventorys = []
        # 基本情報
        inventorys.append(vm.name)                     # VM名
        inventorys.append(vm.config.instanceUuid)      # VMのinstanceUuid
        inventorys.append((lambda x: x.parent.name if(x != None) else "")(vm.resourcePool)) # VMのリソース所属情報
        inventorys.append(vm.guest.toolsStatus)        # VMwareToolsステータス
        inventorys.append(vm.guest.toolsVersion)       # VMwareToolsバージョン
        inventorys.append(vm.config.guestId)           # VMのGuestId

        # CPU情
        inventorys.append(vm.config.hardware.numCoresPerSocket) # CPUソケット数
        inventorys.append(vm.config.hardware.numCPU)            # CPUコア数
        inventorys.append(vm.config.cpuAllocation.reservation)  # CPU予約
        inventorys.append(vm.config.cpuAllocation.limit)        # CPU制限

        # Memory情報
        inventorys.append(vm.config.hardware.memoryMB * 1024 * 1024) # Memory容量
        inventorys.append(vm.config.memoryAllocation.reservation)    # Memory予約
        inventorys.append(vm.config.memoryAllocation.limit)          # Memory制限

        # Virtualデバイス情報の配列
        devices = vm.config.hardware.device

        # Disk以外のドライブ系情報
        vcd = [x.deviceInfo.summary for x in devices if(isinstance(x, vim.vm.device.VirtualCdrom))]
        inventorys.append(vcd[0] if(len(vcd) >= 1) else "")     # CD/DVD マウント情報
        vflpy = [x.deviceInfo.summary for x in devices if(isinstance(x, vim.vm.device.VirtualFloppy))]
        inventorys.append(vflpy[0] if(len(vflpy) >= 1) else "") # Floppy マウント情報
        vusb = [x.deviceInfo.summary for x in devices if(isinstance(x, vim.vm.device.VirtualUSBController))]
        inventorys.append(vusb[0] if(len(vusb) >= 1) else "")   # USB マウント情報

        # Disk情報
        disk_total_size = 0
        for x in devices:
            if(isinstance(x, vim.vm.device.VirtualDisk)):
                disk_total_size = disk_total_size + x.capacityInBytes
        inventorys.append(disk_total_size)   # Disk合計サイズ

        # Network情報
        ipArray = []
        for x in vm.guest.net:
            for ip in x.ipAddress:
                ipArray.append(ip)
        inventorys.append(",".join(sorted(ipArray))) # VMのIPアドレス

        for x in devices:
            if(re.match(r'^Network adapter', x.deviceInfo.label)):
                if(re.match(r'DVSwitch', x.deviceInfo.summary)):
                    inventorys.append(dp_dict[x.backing.port.portgroupKey]) # 分散ポートグループ名
                else:
                    inventorys.append(x.backing.deviceName) # ポートグループ名

        # Inventoryを追加
        (lambda x: all_inventorys.append(x))(inventorys)

    # WSへInventory情報追加
    ws.append(column_name)
    for esxi_inventory in sorted(all_inventorys):
        ws.append(esxi_inventory)

    # Inventoryカラム名セルの塗りつぶし
    fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='629BF7')
    for row in ws:
        # ソート設定
        s = row[0]
        e = row[len(row) - 1]
        ws.auto_filter.ref = "%s%s:%s%s" % (s.column, s.col_idx, e.column, e.col_idx)

        # セルの塗りつぶし
        for cell in row:
            cell.fill = fill
        break

    # WSタブの色設定
    ws.sheet_properties.tabColor = "1072BA"

def create_dp_dict(network_list):
    """
    分散ポートグループの辞書を作成する関数
    仮想マシンオブジェクトの分散ポートグループ情報だと分散ポートグループのキーしか無いため名前がわからない
    そのため、分散ポートグループキーから分散ポートグループ名が取得できるよう辞書を作成する

    :type network_list: class
    :param network_list: ネットワークオブジェクトのリスト

    :rtype: dict
    :return: 分散ポートグループ情報の辞書
    """
    dp_dict = {}
    for nw in network_list.view:
        if(isinstance(nw, vim.dvs.DistributedVirtualPortgroup)):
            dp_dict[nw.config.key] = nw.config.name

    return dp_dict

if __name__ == '__main__':
    # オプションを取得
    args = options()

    # Excel作成
    wb = openpyxl.Workbook()

    # ログイン情報を設定
    login = login()
    login.username = args.username
    login.password = args.password
    login.host = args.host

    # ServiceInstanceを取得
    si = login.get_service_instance()

    # ESXiホストオブジェクト取得
    esxi_list = get_vmware_mob(si, vim.HostSystem)

    # ESXiのInventory作成
    esxi_inventory_ws = wb.active
    esxi_inventory_ws.title = "ESXi Inventory"
    create_esxi_inventorys(esxi_list, esxi_inventory_ws)

    # VMのオブジェクト取得
    vm_list = get_vmware_mob(si, vim.VirtualMachine)

    # 分散ポートグループ情報取得
    network_list = get_vmware_mob(si, vim.Network)
    dp_dict = create_dp_dict(network_list)

    # VMのInventory作成
    vm_inventorys = wb.create_sheet('VM Inventory')
    create_vm_inventorys(vm_list, dp_dict, vm_inventorys)

    # Excel保存
    wb.save(args.output)
